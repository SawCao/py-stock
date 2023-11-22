import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from collections import defaultdict

def get_next(team, point, day, vacation_dict):
    while True:
        real_point = point % len(team)
        if day not in vacation_dict or team[real_point] not in vacation_dict[day]:
            return point
        point += 1
def get_next_emptypoint(team, point, day, vacation_dict, empty_point):
    while True:
        real_point = point % len(team)
        real_empty_point = empty_point % len(team)
        if (day not in vacation_dict or team[real_point] not in vacation_dict[day]) and real_point != real_empty_point:
            return point
        point += 1
# 副班民警
def assign_teams_sub_police(start_date, end_date, police_team, four_team, three_team, two_team, one_team, points, curr_team, vacation_dict):
    start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
    end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
    teams = [police_team, four_team, three_team, two_team, one_team]
    assignments = {}
    days_diff = (end_date - start_date).days + 1
    
    for i in range(days_diff):
        day = (start_date + datetime.timedelta(days=i)).strftime("%Y-%m-%d")
        assignments[day] = []
        real_curr_team = curr_team % len(teams)
        point1 = points[real_curr_team]
        point1 = get_next(teams[real_curr_team], point1, day, vacation_dict)
        point2 = get_next(teams[real_curr_team], point1+1, day, vacation_dict)
        real_point1 = point1 % len(teams[real_curr_team])
        real_point2 = point2 % len(teams[real_curr_team])
        assignments[day].append(teams[real_curr_team][real_point1])
        assignments[day].append(teams[real_curr_team][real_point2])
        if real_point2 < real_point1:
            curr_team += 1
        if real_point2 == len(teams[real_curr_team]) - 1:
            curr_team += 1
        points[real_curr_team] = point2 + 1
    return assignments, points, curr_team

#主班民警
def assign_alone_teams(start_date, end_date, team, point, vacation_dict):
    start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
    end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
    assignments = {}
    days_diff = (end_date - start_date).days + 1

    for i in range(days_diff):
        day = (start_date + datetime.timedelta(days=i)).strftime("%Y-%m-%d")
        assignments[day] = []
        point = get_next(team, point, day, vacation_dict)
        real_point = point % len(team)
        assignments[day].append(team[real_point])
        point += 1
    return assignments, point

#主班领导 完成
def assign_main_leader_teams(start_date, end_date, team, point, vacation_dict):
    start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
    end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
    assignments = {}
    days_diff = (end_date - start_date).days + 1

    for i in range(days_diff):
        day = (start_date + datetime.timedelta(days=i)).strftime("%Y-%m-%d")
        assignments[day] = []
        point = get_next(team, point, day, vacation_dict)
        real_point = point % len(team)
        assignments[day].append(team[real_point])
        point += 1
    return assignments, point

#副班领导 完成
def assign_deputy_leader_teams(start_date, end_date, team, point, empty_point, vacation_dict):
    start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
    end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
    assignments = {}
    days_diff = (end_date - start_date).days + 1
    
    for i in range(days_diff):
        day = (start_date + datetime.timedelta(days=i)).strftime("%Y-%m-%d")
        assignments[day] = []
        start_turn = int(point/len(team))
        point = get_next_emptypoint(team, point, day, vacation_dict, empty_point)
        real_point = point % len(team)
        assignments[day].append(team[real_point])
        point += 1
        if int(point/len(team)) > start_turn:
            empty_point += 1
    return assignments, point


def output_to_existing_excel(assignments, filename, first_col, second_col, interval):
    # 加载现有的 Excel 文件
    wb = load_workbook(filename)
    ws = wb.active
    
    # 初始位置
    col_start = {first_col: 3, second_col: 3}
    
    # 当前位置
    col_pos = {first_col: 3, second_col: 3}

    # 每列日期的最大条目数
    max_items_per_col = 15

    date_column = first_col

    date_count = 0
    for date, members in assignments.items():
        # 添加日期
        ws[f'{date_column}{col_pos[date_column]}'] = date

        # 添加成员
        members_str = '\n'.join(members)
        ws[f'{get_column_letter(ws[f"{date_column}{col_pos[date_column]}"].column + interval)}{col_pos[date_column]}'] = members_str

        # 更新位置
        col_pos[date_column] += 1
        date_count += 1
        # 检查是否需要切换到另一列
        if date_count >= max_items_per_col:
            date_count = 0;
            date_column = second_col

    # 保存修改后的 Excel 文件
    wb.save(filename)

def output_to_existing_excel_main_police(assignments, filename):
    # 加载现有的 Excel 文件
    wb = load_workbook(filename)
    ws = wb.active
    
    # 初始位置
    col_start = {'A': 3, 'D': 3}
    
    # 当前位置
    col_pos = {'A': 3, 'D': 3}

    # 每列日期的最大条目数
    max_items_per_col = 15

    date_column = 'A'
    date_count = 0
    for date, members in assignments.items():
        # 添加日期
        ws[f'{date_column}{col_pos[date_column]}'] = date

        # 添加成员
        members_str = '\n'.join(members)
        ws[f'{get_column_letter(ws[f"{date_column}{col_pos[date_column]}"].column + 1)}{col_pos[date_column]}'] = members_str

        # 更新位置
        col_pos[date_column] += 1
        date_count += 1
        # 检查是否需要切换到另一列
        if date_count >= max_items_per_col:
            date_count = 0;
            date_column = 'D'

    # 保存修改后的 Excel 文件
    wb.save(filename)

# 一中队
zhongdui_1 = [
    "李楷锋13068998802", "杨坚凯15817930211", 
    "王育彬13342740810", "黄春山13924789065",
    "翁培嘉18823991992", "陈汉强15876154500", 
    "林荣毅15876172597", "刘涛13719921969",
    "林烁13025224218", "罗松滨15817942205", 
    "陈祺涛18998249232", "连益民15916647050",
    "赵嘉诺13288001655", "林晓丹13192370536"
]

# 二中队
zhongdui_2 = [
    "余昌元15815015018", "张锦杰15815091983",
    "张海斌16607549896", "林怡荣15815102866",
    "余斌强16626101117", "赖栋15816617195",
    "许荣腾13682882743", "廖继乾13546824415",
    "谢振涛13428333019", "林喜镇15889213774",
    "林源15989845432", "翁建平15019725014",
    "余泽斌15916622787", "陈鳅洋13501414442"
]

# 三中队
zhongdui_3 = [
    "胡煜15815068798", "林少伟15019706212",
    "唐惟识13531163344", "吴烨源15019778688",
    "林泽佳15992281749", "吴承焓13433828298",
    "陈少伟13415198944", "范玮欣15875464046",
    "黄钦峰13016660014", "黄腾13428302081",
    "许晓章13825889679", "吴坤宏13411931755",
    "陈泽斌13414073644"
]

# 四中队（无）

# 五中队
zhongdui_5 = [
    "翁介南18029558755", "陈程鹏15113117722",
    "詹世鑫13226816365", "谢思凯13539656383",
    "陈杰彬13794104401", "朱延鑫13682875298",
    "林俊鑫17607549713", "陈华佳13226827920",
    "蔡国森13682877737", "陈佳豪13433854018"
]

# 骑警中队
zhongdui_qijing = [
    "陈胜洲13502997495", "庄锦坤15815000994", 
    "谢燕彬13028187978", "王金湖13556450433", 
    "郑裕亮13829691586", "邱源18675455859", 
    "方捷15816615469", "方鸿18664464272",
    "高涛18811125186", "刘灵驱15889251774", 
    "陈宇泓13113491860", "高塬辉13246082668",
    "黄少涵13428314449", "吴绍吉13428319568", 
    "刘拥军15626170755"
]

leadership = [
    ["文伟钿", "蔡应谦", "范锦炎", "丁银才", "马立光", "宗广荣"],
    ["杨伟", "李伟嘉", "林国胜", "吴钦辉", "林经智", "林瑞锋", "郑康弟"],
    ["林奕兵", "周文迅", "洪永海", "王彪", "林少程", "许夏阳", "邱华先", "徐卓辉", "林建辉", "张少卫", "郑文斌", "高泽波", "肖建全", "李耘", "陈旭彬", "许小刚", "杨少鸿", "林静山", "张仲凯", "陈敬斌", "马少雄", "陈雄文", "吕俊伟", "林敬钦", "朱少辉", "郑桂峰", "吴贤才", "黄献忠", "李业雄", "陈佳建", "王江勇"],
    ["王江勇", "陈佳建", "李业雄", "黄献忠", "吴贤才", "郑桂峰", "朱少辉", "林敬钦", "吕俊伟", "陈雄文", "马少雄", "陈敬斌", "张仲凯", "林静山", "杨少鸿", "许小刚", "陈旭彬", "李耘", "肖建全", "高泽波", "郑文斌", "张少卫", "林建辉", "徐卓辉", "邱华先", "许夏阳", "林少程", "王彪", "洪永海", "周文迅", "林奕兵"]
]

vacation_dict= {
    "2023-01-01":['李楷锋13068998802','文伟钿','杨伟'],
}
# 测试
# 安保测试
start_date = "2023-01-01"
end_date = "2023-01-30"
# result1, point, curr_team = assign_teams_sub_police(start_date, end_date, zhongdui_1, zhongdui_2, [0,0], 0, vacation_dict)
# start_date = "2023-02-01"
# end_date = "2023-02-28"
# result2, point, curr_team = assign_teams_sub_police(start_date, end_date, zhongdui_1, zhongdui_2, point, curr_team, vacation_dict)
# start_date = "2023-03-01"
# end_date = "2023-03-31"
# result3, point, curr_team = assign_teams_sub_police(start_date, end_date, zhongdui_1, zhongdui_2, point, curr_team, vacation_dict)
# for day, assigned in result1.items():
#     print(f"{day}: {assigned}")
# for day, assigned in result2.items():
#     print(f"{day}: {assigned}")
# for day, assigned in result3.items():
#     print(f"{day}: {assigned}")
    
    
leader_result, point = assign_main_leader_teams(start_date, end_date, leadership[0], 0, vacation_dict)
output_to_existing_excel(leader_result, "two.xlsx", 'A', 'F', 1)
deputy_leader_teams, main_police_point = assign_deputy_leader_teams(start_date, end_date, leadership[1], 0, 0, vacation_dict )
output_to_existing_excel(deputy_leader_teams, "two.xlsx", 'A', 'F', 2)
Main_Police, main_police_point = assign_alone_teams(start_date, end_date, leadership[2], 0, vacation_dict)
output_to_existing_excel(Main_Police, "two.xlsx", 'A', 'F',3 )
sub_police, point, curr_team = assign_teams_sub_police(start_date, end_date, zhongdui_1, zhongdui_2, zhongdui_3, zhongdui_5, zhongdui_qijing, [0,0,0,0,0], 0, vacation_dict)
deputy_Police, main_police_point = assign_alone_teams(start_date, end_date, leadership[3], 0, vacation_dict)
# 定义一个空字典作为结果
subanddeputy_police = {}

# 将dict1和dict2中的键值对加入到merged_dict对象中
for key, value in deputy_Police.items():
    subanddeputy_police.setdefault(key, []).extend(value)
for key, value in sub_police.items():
    subanddeputy_police.setdefault(key, []).extend(value)
subanddeputy_police = {key: values for key, values in subanddeputy_police.items() if len(values) > 1}
output_to_existing_excel(subanddeputy_police, "two.xlsx", 'A', 'F', 4)
# # deputy_leader_teams, main_police_point = assign_deputy_leader_teams(start_date, end_date, leadership[1], 0, 0, vacation_dict )
# # for day, assigned in deputy_leader_teams.items():
# #     print(f"{day}: {assigned}")
# # 调用函数
# output_to_existing_excel(result1, "one.xlsx")
# output_to_existing_excel_main_police(Main_Police, "one.xlsx")


# result2, points, curr_team, next_team = assign_teams(start_date, end_date, police_team, four_team, three_team, two_team, one_team, points, curr_team, next_team)
# for day, assigned in result2.items():
#     print(f"2--{day}: {assigned}")