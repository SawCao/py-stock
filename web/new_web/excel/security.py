import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import calendar

class securityExcel:
    def __init__(self, start_date, end_date, leaderships, sub_teams, vacation_dict, main_police_point, security_points, security_curr_team, security_next_team):
        self.start_date = start_date
        self.end_date = end_date
        self.leaderships = leaderships
        self.sub_teams = sub_teams
        self.vacation_dict = vacation_dict
        self.main_police_point = main_police_point
        self.security_points = security_points
        self.security_curr_team = security_curr_team
        self.security_next_team = security_next_team

    def get_next(self, team, point, day, vacation_dict):
        while True:
            real_point = point % len(team)
            if day not in vacation_dict or team[real_point] not in vacation_dict[day]:
                return point
            point += 1
            
    def assign_teams(self):
        start_date = datetime.datetime.strptime(self.start_date, "%Y-%m-%d")
        end_date = datetime.datetime.strptime(self.end_date, "%Y-%m-%d")
        teams = self.sub_teams
        points = self.security_points
        curr_team = self.security_curr_team
        next_team = self.security_next_team
        assignments = {}
        days_diff = (end_date - start_date).days + 1


        for i in range(days_diff):
            day = (start_date + datetime.timedelta(days=i)).strftime("%Y-%m-%d")
            assignments[day] = []
            real_curr_team = curr_team % len(teams)
            real_next_team = next_team % len(teams)
            point1 = points[real_curr_team]
            point2 = points[real_next_team]
            point1 = self.get_next(teams[real_curr_team], point1, day, vacation_dict)
            point2 = self.get_next(teams[real_next_team], point2, day, vacation_dict)
            real_point1 = point1 % len(teams[real_curr_team])
            real_point2 = point2 % len(teams[real_next_team])
            assignments[day].append(teams[real_curr_team][real_point1])
            assignments[day].append(teams[real_next_team][real_point2])
            points[real_curr_team] = point1 + 1
            points[real_next_team] = point2 + 1
            if real_point1 == len(teams[real_curr_team]) - 1:
                curr_team = next_team
                next_team = curr_team + 1
            if real_point2 == len(teams[real_next_team]) - 1:
                next_team += 1
        self.security_points = points
        self.security_next_team = next_team
        self.security_curr_team = curr_team
        return assignments

    #主班民警
    def assign_alone_teams(self):
        start_date = datetime.datetime.strptime(self.start_date, "%Y-%m-%d")
        end_date = datetime.datetime.strptime(self.end_date, "%Y-%m-%d")
        assignments = {}
        days_diff = (end_date - start_date).days + 1
        team = self.leaderships[2]
        point = self.main_police_point
        for i in range(days_diff):
            day = (start_date + datetime.timedelta(days=i)).strftime("%Y-%m-%d")
            assignments[day] = []
            point = self.get_next(team, point, day, vacation_dict)
            real_point = point % len(team)
            assignments[day].append(team[real_point])
            point += 1
        self.main_police_point = point
        return assignments
    
    def output_to_existing_excel(self, assignments, filename, first_col, second_col, interval, save_filename):
        # 加载现有 Excel 文件，并获取其活动工作表
        wb = load_workbook(filename)
        ws = wb.active

        # 获取 assignments 中第一个日期并格式化为 "%Y年%m月"
        first_date = datetime.datetime.strptime(list(assignments.keys())[0], '%Y-%m-%d').strftime('%Y年%m月')

        # 更新 A1 单元格的值
        ws['A1'] = f'交警机动巡逻大队{first_date}大楼安全保卫安排表'

        # 初始位置
        col_start = {first_col: 3, second_col: 3}

        # 当前位置
        col_pos = {first_col: 3, second_col: 3}

        # 每列日期的最大条目数
        max_items_per_col = 15

        date_column = first_col

        date_count = 0
        for date, members in assignments.items():
            # 添加日期和星期信息
            weekday = calendar.day_name[datetime.datetime.strptime(date, '%Y-%m-%d').weekday()]
            weekday_cn = "一二三四五六日"[datetime.datetime.strptime(date, '%Y-%m-%d').weekday()]

            ws[f'{date_column}{col_pos[date_column]}'] = f'{date}\n星期{weekday_cn}'

            # 添加成员
            members_str = '\n'.join(members)
            ws[f'{get_column_letter(ws[f"{date_column}{col_pos[date_column]}"].column + interval)}{col_pos[date_column]}'] = members_str

            # 更新位置
            col_pos[date_column] += 1
            date_count += 1
            
            # 检查是否需要切换到另一列
            if date_count >= max_items_per_col:
                date_count = 0
                date_column = second_col

        # 保存修改后的 Excel 文件
        wb.save(save_filename)
    def RUN(self):
        result1 = self.assign_teams()
        Main_Police = self.assign_alone_teams()
        # 调用函数
        self.output_to_existing_excel(result1, "./static/spreadsheets/one_empty.xlsx", "A", "D", 2, "./static/spreadsheets/one.xlsx")
        self.output_to_existing_excel(Main_Police, "./static/spreadsheets/one.xlsx","A", "D", 1, "./static/spreadsheets/one.xlsx")

        for day, assigned in result1.items():
            print(f"{day}: {assigned}")
        for day, assigned in Main_Police.items():
            print(f"{day}: {assigned}")

# 一中队名单
one_team = [
    '林晓丹13192370536', '赵嘉诺13288001655', '连益民15916647050',
    '陈祺涛18998249232', '罗松滨15817942205', '林烁13025224218',
    '刘涛13719921969', '林荣毅15876172597', '陈汉强15876154500',
    '翁培嘉18823991992', '黄春山13924789065', '王育彬13342740810',
    '杨坚凯15817930211', '李楷锋13068998802'
]

# 二中队名单
two_team = [
    '陈鳅洋13501414442', '余泽斌15916622787', '翁建平15019725014',
    '林源15989845432', '林喜镇15889213774', '谢振涛13428333019',
    '廖继乾13546824415', '许荣腾13682882743', '赖栋15816617195',
    '余斌强16626101117', '林怡荣15815102866', '张海斌16607549896',
    '张锦杰15815091983', '余昌元15815015018'
]

# 三中队名单
three_team = [
    '陈泽斌13414073644', '吴坤宏13411931755', '许晓章13825889679',
    '黄腾13428302081', '黄钦峰13016660014', '范玮欣15875464046',
    '陈少伟13415198944', '吴承焓13433828298', '林泽佳15992281749',
    '吴烨源15019778688', '唐惟识13531163344', '林少伟15019706212',
    '胡煜15815068798'
]

# 四中队名单
four_team = [
    '翁煜斌13501408777', '陈佳豪13433854018', '蔡国森13682877737',
    '陈华佳13226827920', '林俊鑫17607549713', '林伟涛18650198081',
    '朱延鑫13682875298', '陈杰彬13794104401', '谢思凯13539656383',
    '詹世鑫13226816365', '陈程鹏15113117722', '翁介南18029558755',
    '杨宇泓18165904036'
]

# 骑警中队名单
police_team = [
    '刘拥军15626170755', '吴绍吉13428319568', '黄少涵13428314449',
    '高塬辉13246082668', '陈宇泓13113491860', '刘灵驱15889251774',
    '高涛18811125186', '方鸿18664464272', '方捷15816615469',
    '邱源18675455859', '郑裕亮13829691586', '王金湖13556450433',
    '谢燕彬13028187978', '庄锦坤15815000994', '陈胜洲13502997495'
]

leadership = [
    ["文伟钿", "蔡应谦", "范锦炎", "丁银才", "马立光", "宗广荣"],
    ["杨伟", "李伟嘉", "林国胜", "吴钦辉", "林经智", "林瑞锋", "郑康弟"],
    ["林奕兵", "周文迅", "洪永海", "王彪", "林少程", "许夏阳", "邱华先", "徐卓辉", "林建辉", "张少卫", "郑文斌", "高泽波", "肖建全", "李耘", "陈旭彬", "许小刚", "杨少鸿", "林静山", "张仲凯", "陈敬斌", "马少雄", "陈雄文", "吕俊伟", "林敬钦", "朱少辉", "郑桂峰", "吴贤才", "黄献忠", "李业雄", "陈佳建", "王江勇"],
    ["王江勇", "陈佳建", "李业雄", "黄献忠", "吴贤才", "郑桂峰", "朱少辉", "林敬钦", "吕俊伟", "陈雄文", "马少雄", "陈敬斌", "张仲凯", "林静山", "杨少鸿", "许小刚", "陈旭彬", "李耘", "肖建全", "高泽波", "郑文斌", "张少卫", "林建辉", "徐卓辉", "邱华先", "许夏阳", "林少程", "王彪", "洪永海", "周文迅", "林奕兵"]
]

vacation_dict= {
    "2023-01-01":['文伟钿','杨伟','刘拥军15626170755'],
}

# 测试
start_date = "2023-01-01"
end_date = "2023-01-30"

# subteam = [police_team, four_team, three_team, two_team, one_team]
# oncall = securityExcel(start_date, end_date, leadership, subteam, vacation_dict, 0, [0,0,0,0,0], 0, 1)
# oncall.RUN()