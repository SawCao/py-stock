import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def assign_teams(start_date, end_date, police_team, four_team, three_team, two_team, one_team, points, curr_team, next_team):
    start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d")
    end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d")
    teams = [police_team, four_team, three_team, two_team, one_team]
    assignments = {}
    days_diff = (end_date - start_date).days + 1


    for i in range(days_diff):
        day = (start_date + datetime.timedelta(days=i)).strftime("%Y-%m-%d")
        assignments[day] = []

        if points[curr_team] <= len(teams[curr_team]) - 1:
            first_people = teams[curr_team][points[curr_team]]
        else:
            points[curr_team] = 0
            curr_team = next_team
            next_team = (curr_team + 1) % len(teams)
            first_people = teams[curr_team][points[curr_team]]
        points[curr_team] += 1
        
        
        if points[next_team] <= len(teams[next_team]) - 1:
            second_people = teams[next_team][points[next_team]]
        else:
            points[next_team] = 0
            next_team +=1
            second_people = teams[next_team][points[next_team]]
        points[next_team] += 1
        assignments[day].append(first_people)
        assignments[day].append(second_people) 
    return assignments, points, curr_team, next_team

def output_to_excel(assignments):
    wb = Workbook()
    ws = wb.active

    # 初始位置
    col_start = {'A': 3, 'F': 3}
    
    # 当前位置
    col_pos = {'A': 3, 'F': 3}

    # 每列日期的最大条目数
    max_items_per_col = 15

    date_column = 'A'
    date_count = 0
    for date, members in assignments.items():
        # 添加日期
        ws[f'{date_column}{col_pos[date_column]}'] = date

        # 添加成员
        members_str = '，'.join(members)
        ws[f'{get_column_letter(ws[f"{date_column}{col_pos[date_column]}"].column + 1)}{col_pos[date_column]}'] = members_str

        # 更新位置
        col_pos[date_column] += 1
        date_count += 1
        # 检查是否需要切换到另一列
        if date_count >= max_items_per_col:
            date_count = 0
            date_column = 'F'

    # 保存 Excel 文件
    wb.save("assignments.xlsx")
    
# 一中队
zhongdui_1 = [
    "李楷锋13068998802", "杨坚凯15817930211", 
    "王育彬13342740810", "黄春山13924789065",
    "翁培嘉18823991992", "陈汉强15876154500", 
    "林荣毅15876172597", "刘涛13719921969",
    "林烁13025224218", "罗松滨15817942205", 
    "陈祺涛18998249232", "连益民15916647050",
    "赵嘉诺13288001655", "林晓丹13192370536"
]

# 二中队
zhongdui_2 = [
    "余昌元15815015018", "张锦杰15815091983",
    "张海斌16607549896", "林怡荣15815102866",
    "余斌强16626101117", "赖栋15816617195",
    "许荣腾13682882743", "廖继乾13546824415",
    "谢振涛13428333019", "林喜镇15889213774",
    "林源15989845432", "翁建平15019725014",
    "余泽斌15916622787", "陈鳅洋13501414442"
]

# 三中队
zhongdui_3 = [
    "胡煜15815068798", "林少伟15019706212",
    "唐惟识13531163344", "吴烨源15019778688",
    "林泽佳15992281749", "吴承焓13433828298",
    "陈少伟13415198944", "范玮欣15875464046",
    "黄钦峰13016660014", "黄腾13428302081",
    "许晓章13825889679", "吴坤宏13411931755",
    "陈泽斌13414073644"
]

# 四中队（无）

# 五中队
zhongdui_5 = [
    "翁介南18029558755", "陈程鹏15113117722",
    "詹世鑫13226816365", "谢思凯13539656383",
    "陈杰彬13794104401", "朱延鑫13682875298",
    "林俊鑫17607549713", "陈华佳13226827920",
    "蔡国森13682877737", "陈佳豪13433854018"
]

# 骑警中队
zhongdui_qijing = [
    "陈胜洲13502997495", "庄锦坤15815000994", 
    "谢燕彬13028187978", "王金湖13556450433", 
    "郑裕亮13829691586", "邱源18675455859", 
    "方捷15816615469", "方鸿18664464272",
    "高涛18811125186", "刘灵驱15889251774", 
    "陈宇泓13113491860", "高塬辉13246082668",
    "黄少涵13428314449", "吴绍吉13428319568", 
    "刘拥军15626170755"
]

# 测试
start_date = "2023-01-01"
end_date = "2023-01-30"
result1, points, curr_team, next_team = assign_teams(start_date, end_date, zhongdui_1, zhongdui_2, zhongdui_3, zhongdui_5, zhongdui_qijing, 0, 0, [0,0,0,0,0])
output_to_excel(result1)
for day, assigned in result1.items():
    print(f"{day}: {assigned}")
result2, points, curr_team, next_team = assign_teams(start_date, end_date, police_team, four_team, three_team, two_team, one_team, points, curr_team, next_team)
for day, assigned in result2.items():
    print(f"2--{day}: {assigned}")